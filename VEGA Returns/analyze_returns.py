import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# Read the monthly returns data
df = pd.read_csv('VEGA_Monthly_Returns.csv')

# Starting capital
INITIAL_CAPITAL = 1_000_000
RISK_FREE_RATE = 0.0  # Assuming 0% for simplicity, can adjust

# Calculate monthly returns based on beginning equity each month
# Return % should be calculated from the previous month's equity
df['Month'] = pd.to_datetime(df['Month'])
df = df.sort_values('Month').reset_index(drop=True)

# Recalculate returns properly - return is P&L / Beginning Equity
# Beginning equity for month N = Equity at end of month N-1
df['Beginning_Equity'] = df['Equity'].shift(1)
df.loc[0, 'Beginning_Equity'] = INITIAL_CAPITAL
df['Monthly_Return'] = df['Monthly_PnL'] / df['Beginning_Equity']

# Fill in months with 0 trades (they exist in the data with 0 P&L)
# The data already accounts for all months

# --- Calculate Statistics ---

# Monthly returns array
monthly_returns = df['Monthly_Return'].values

# Average Monthly Return
avg_monthly_return = np.mean(monthly_returns)

# Average Annual Return (compounded)
total_return = df['Equity'].iloc[-1] / INITIAL_CAPITAL - 1
n_years = len(df) / 12
cagr = (1 + total_return) ** (1 / n_years) - 1

# Simple annualized return (avg monthly * 12)
avg_annual_return_simple = avg_monthly_return * 12

# Monthly Standard Deviation
monthly_std = np.std(monthly_returns, ddof=1)

# Annual Standard Deviation
annual_std = monthly_std * np.sqrt(12)

# Sharpe Ratio (using monthly data, then annualized)
# Sharpe = (Avg Return - Risk Free) / Std Dev
monthly_rf = RISK_FREE_RATE / 12
sharpe_monthly = (avg_monthly_return - monthly_rf) / monthly_std
sharpe_annual = sharpe_monthly * np.sqrt(12)

# Sortino Ratio (using only downside deviation)
negative_returns = monthly_returns[monthly_returns < 0]
downside_std = np.std(negative_returns, ddof=1) if len(negative_returns) > 1 else 0
sortino_monthly = (avg_monthly_return - monthly_rf) / downside_std if downside_std > 0 else np.nan
sortino_annual = sortino_monthly * np.sqrt(12) if not np.isnan(sortino_monthly) else np.nan

# Maximum Drawdown (peak to valley)
cumulative_equity = df['Equity'].values
running_max = np.maximum.accumulate(cumulative_equity)
drawdowns = (cumulative_equity - running_max) / running_max
max_drawdown = np.min(drawdowns)
max_drawdown_abs = np.min(cumulative_equity - running_max)

# Calmar Ratio (CAGR / |Max Drawdown|)
calmar_ratio = cagr / abs(max_drawdown) if max_drawdown != 0 else np.nan

# Trade Statistics (from monthly data)
total_trades = df['Trades'].sum()
total_winners = df['Winners'].sum()
total_losers = df['Losers'].sum()

# Calculate average winner/loser from the individual strategy files
# We need to aggregate from the CSVs
strategies = [
    ('VEGA LE DN NI.csv', 'LE DN NI'),
    ('VEGA LE UP NI.csv', 'LE UP NI'),
    ('VEGA SE DN NI.csv', 'SE DN NI'),
    ('VEGA SE UP NI.csv', 'SE UP NI')
]

# Parse strategy data to get gross profit/loss
total_gross_profit = 0
total_gross_loss = 0
total_winning_trades = 0
total_losing_trades = 0

for filename, name in strategies:
    try:
        with open(filename, 'r') as f:
            content = f.read()

        # Extract Gross Profit
        for line in content.split('\n'):
            if line.startswith('Gross Profit,'):
                parts = line.split(',')
                val = parts[1].replace('$', '').replace(',', '').replace('(', '-').replace(')', '')
                total_gross_profit += float(val)
            elif line.startswith('Gross Loss,'):
                parts = line.split(',')
                val = parts[1].replace('$', '').replace(',', '').replace('(', '-').replace(')', '')
                total_gross_loss += abs(float(val))
            elif line.startswith('Winning Trades,'):
                parts = line.split(',')
                total_winning_trades += int(parts[1])
            elif line.startswith('Losing Trades,'):
                parts = line.split(',')
                total_losing_trades += int(parts[1])
    except Exception as e:
        print(f"Error reading {filename}: {e}")

avg_winner = total_gross_profit / total_winning_trades if total_winning_trades > 0 else 0
avg_loser = -total_gross_loss / total_losing_trades if total_losing_trades > 0 else 0
avg_net = (total_gross_profit - total_gross_loss) / (total_winning_trades + total_losing_trades) if (total_winning_trades + total_losing_trades) > 0 else 0

# --- Create Excel Output ---
wb = Workbook()
ws = wb.active
ws.title = "Performance Summary"

# Styling
header_font = Font(bold=True, size=12)
title_font = Font(bold=True, size=14)
money_format = '#,##0.00'
pct_format = '0.00%'
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font_white = Font(bold=True, color="FFFFFF")

# Title
ws['A1'] = "VEGA Combined Strategy Performance Analysis"
ws['A1'].font = title_font
ws.merge_cells('A1:D1')

# Statistics Section
stats_start = 3
ws[f'A{stats_start}'] = "Performance Statistics"
ws[f'A{stats_start}'].font = header_font

stats = [
    ("Average Monthly Return", f"{avg_monthly_return:.4%}"),
    ("Average Annual Return (CAGR)", f"{cagr:.4%}"),
    ("Monthly Standard Deviation", f"{monthly_std:.4%}"),
    ("Annual Standard Deviation", f"{annual_std:.4%}"),
    ("Sharpe Ratio (Annualized)", f"{sharpe_annual:.2f}"),
    ("Sortino Ratio (Annualized)", f"{sortino_annual:.2f}"),
    ("Calmar Ratio", f"{calmar_ratio:.2f}"),
    ("Maximum Drawdown", f"{max_drawdown:.4%}"),
    ("Maximum Drawdown ($)", f"${max_drawdown_abs:,.2f}"),
    ("", ""),
    ("Trade Statistics", ""),
    ("Total Trades", f"{int(total_trades)}"),
    ("Winning Trades", f"{int(total_winners)}"),
    ("Losing Trades", f"{int(total_losers)}"),
    ("Win Rate", f"{total_winners/total_trades:.2%}" if total_trades > 0 else "N/A"),
    ("Average Winner", f"${avg_winner:,.2f}"),
    ("Average Loser", f"${avg_loser:,.2f}"),
    ("Average Net Trade", f"${avg_net:,.2f}"),
]

for i, (label, value) in enumerate(stats):
    row = stats_start + 1 + i
    ws[f'A{row}'] = label
    ws[f'B{row}'] = value
    if label in ["Performance Statistics", "Trade Statistics"]:
        ws[f'A{row}'].font = header_font

# Monthly Returns Section
returns_start = stats_start + len(stats) + 3
ws[f'A{returns_start}'] = "Monthly Returns"
ws[f'A{returns_start}'].font = header_font

# Headers
headers = ['Month', 'P&L ($)', 'Return (%)', 'Cumulative P&L', 'Equity', 'Trades', 'Winners', 'Losers']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=returns_start + 1, column=col, value=header)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center')

# Data
for idx, row in df.iterrows():
    data_row = returns_start + 2 + idx
    ws.cell(row=data_row, column=1, value=row['Month'].strftime('%Y-%m'))
    ws.cell(row=data_row, column=2, value=row['Monthly_PnL'])
    ws.cell(row=data_row, column=3, value=row['Monthly_Return'])
    ws.cell(row=data_row, column=4, value=row['Cumulative_PnL'])
    ws.cell(row=data_row, column=5, value=row['Equity'])
    ws.cell(row=data_row, column=6, value=row['Trades'])
    ws.cell(row=data_row, column=7, value=row['Winners'])
    ws.cell(row=data_row, column=8, value=row['Losers'])

    # Format cells
    ws.cell(row=data_row, column=2).number_format = money_format
    ws.cell(row=data_row, column=3).number_format = pct_format
    ws.cell(row=data_row, column=4).number_format = money_format
    ws.cell(row=data_row, column=5).number_format = money_format

    for col in range(1, 9):
        ws.cell(row=data_row, column=col).border = thin_border

# Adjust column widths
ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 15
ws.column_dimensions['F'].width = 10
ws.column_dimensions['G'].width = 10
ws.column_dimensions['H'].width = 10

# Create Chart
chart = LineChart()
chart.title = "Monthly Returns (%)"
chart.style = 10
chart.y_axis.title = "Return %"
chart.x_axis.title = "Month"
chart.height = 12
chart.width = 20

# Data for chart
data_end_row = returns_start + 1 + len(df)
data = Reference(ws, min_col=3, min_row=returns_start + 1, max_row=data_end_row)
cats = Reference(ws, min_col=1, min_row=returns_start + 2, max_row=data_end_row)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)

# Position chart
chart_row = returns_start + len(df) + 5
ws.add_chart(chart, f"A{chart_row}")

# Also add an equity curve chart
equity_chart = LineChart()
equity_chart.title = "Equity Curve"
equity_chart.style = 10
equity_chart.y_axis.title = "Equity ($)"
equity_chart.x_axis.title = "Month"
equity_chart.height = 12
equity_chart.width = 20

equity_data = Reference(ws, min_col=5, min_row=returns_start + 1, max_row=data_end_row)
equity_chart.add_data(equity_data, titles_from_data=True)
equity_chart.set_categories(cats)

ws.add_chart(equity_chart, f"K{chart_row}")

# Save
output_file = 'VEGA_Performance_Analysis_Combined.xlsx'
wb.save(output_file)
print(f"Analysis saved to {output_file}")

# Print summary to console
print("\n" + "="*60)
print("VEGA Combined Strategy Performance Summary")
print("="*60)
print(f"\nAverage Monthly Return:        {avg_monthly_return:.4%}")
print(f"Average Annual Return (CAGR):  {cagr:.4%}")
print(f"Monthly Standard Deviation:    {monthly_std:.4%}")
print(f"Annual Standard Deviation:     {annual_std:.4%}")
print(f"\nSharpe Ratio (Annualized):     {sharpe_annual:.2f}")
print(f"Sortino Ratio (Annualized):    {sortino_annual:.2f}")
print(f"Calmar Ratio:                  {calmar_ratio:.2f}")
print(f"\nMaximum Drawdown:              {max_drawdown:.4%}")
print(f"Maximum Drawdown ($):          ${max_drawdown_abs:,.2f}")
print(f"\n--- Trade Statistics ---")
print(f"Total Trades:                  {int(total_trades)}")
print(f"Winning Trades:                {int(total_winners)}")
print(f"Losing Trades:                 {int(total_losers)}")
print(f"Win Rate:                      {total_winners/total_trades:.2%}")
print(f"\nAverage Winner:                ${avg_winner:,.2f}")
print(f"Average Loser:                 ${avg_loser:,.2f}")
print(f"Average Net Trade:             ${avg_net:,.2f}")
print("="*60)
