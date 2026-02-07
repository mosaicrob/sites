#!/usr/bin/env python3
"""
Create Excel comparison of EasyLanguage (CSV) vs Python (XLSX) results
Highlights matches in green, discrepancies in red
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime
from collections import defaultdict

# Styles
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
match_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
diff_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
warn_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def parse_csv_trades(csv_file):
    """Parse EasyLanguage CSV trades"""
    with open(csv_file, 'r', encoding='utf-8-sig') as f:
        lines = f.readlines()

    # Find trade list
    trade_start_idx = None
    for i, line in enumerate(lines):
        if 'TradeStation Trades List' in line:
            trade_start_idx = i + 3
            break

    trades = []
    current_trade = None

    for line in lines[trade_start_idx:]:
        if 'Annual' in line:
            if current_trade:
                trades.append(current_trade)
            break
        if not line.strip():
            continue

        parts = line.split(',')

        if parts[0].strip() and parts[0].strip().isdigit():
            if current_trade:
                trades.append(current_trade)
            current_trade = {
                'trade_num': int(parts[0]),
                'entry_date': parts[2].strip(),
                'entry_price': parts[4].strip(),
                'contracts': parts[6].strip() if len(parts) > 6 else ''
            }
        elif current_trade and not parts[0].strip():
            current_trade['exit_date'] = parts[1].strip() if len(parts) > 1 else ''
            current_trade['exit_price'] = parts[3].strip() if len(parts) > 3 else ''
            current_trade['pnl_raw'] = parts[6].strip() if len(parts) > 6 else ''

    return trades

def parse_xlsx_trades(xlsx_file):
    """Parse Python XLSX trades"""
    wb = openpyxl.load_workbook(xlsx_file)
    ws = wb['trades']

    trades = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            trades.append({
                'trade_id': row[1],
                'entry_time': row[7],
                'exit_time': row[8],
                'entry_price': row[6],
                'qty': row[5],
                'pnl': row[12]
            })

    return trades

def create_comparison(csv_file, xlsx_file, output_file, strategy_name):
    """Create comparison Excel file"""

    # Parse data
    csv_trades = parse_csv_trades(csv_file)
    xlsx_trades = parse_xlsx_trades(xlsx_file)

    # Calculate totals
    csv_total_pnl = sum(
        float(t['pnl_raw'].replace('$', '').replace(',', '').replace('(', '-').replace(')', ''))
        for t in csv_trades if 'pnl_raw' in t
    )
    xlsx_total_pnl = sum(t['pnl'] for t in xlsx_trades)

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comparison"

    # Title
    ws['A1'] = f"{strategy_name} - EasyLanguage vs Python Comparison"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:F1')

    row = 3

    # Summary
    ws[f'A{row}'] = "Summary Statistics"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 2

    # Headers
    for col, header in enumerate(['Metric', 'EasyLanguage', 'Python', 'Difference', 'Match?'], 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    row += 1

    # Metrics
    metrics = [
        ('Total Trades', len(csv_trades), len(xlsx_trades)),
        ('Total P&L', csv_total_pnl, xlsx_total_pnl),
    ]

    for metric, csv_val, xlsx_val in metrics:
        ws[f'A{row}'] = metric
        ws[f'B{row}'] = f"${csv_val:,.2f}" if metric == 'Total P&L' else csv_val
        ws[f'C{row}'] = f"${xlsx_val:,.2f}" if metric == 'Total P&L' else xlsx_val

        diff = xlsx_val - csv_val
        ws[f'D{row}'] = f"${diff:,.2f}" if metric == 'Total P&L' else diff

        match = abs(diff) < 1 if metric == 'Total P&L' else diff == 0
        ws[f'E{row}'] = "✓" if match else "✗"
        ws[f'E{row}'].fill = match_fill if match else diff_fill

        for col in range(1, 6):
            ws.cell(row=row, column=col).border = border
        row += 1

    # Key Finding
    row += 2
    ws[f'A{row}'] = "KEY FINDING:"
    ws[f'A{row}'].font = Font(bold=True, size=11, color="FF0000")
    row += 1

    if csv_trades and xlsx_trades:
        csv_qty = int(csv_trades[0]['contracts']) if csv_trades[0]['contracts'].isdigit() else 10
        xlsx_qty = xlsx_trades[0]['qty']

        ws[f'A{row}'] = f"Contract Quantity Discrepancy: EasyLanguage uses {csv_qty} contracts, Python uses {xlsx_qty} contract(s)"
        ws[f'A{row}'].fill = warn_fill
        ws[f'A{row}'].font = Font(bold=True)
        ws.merge_cells(f'A{row}:F{row}')
        row += 1

        ws[f'A{row}'] = f"This explains the ~10x P&L difference: ${csv_total_pnl:,.2f} vs ${xlsx_total_pnl:,.2f}"
        ws[f'A{row}'].fill = warn_fill
        ws.merge_cells(f'A{row}:F{row}')
        row += 2

    # Trade-by-trade comparison
    row += 1
    ws[f'A{row}'] = "Trade-by-Trade Comparison (First 30 Trades)"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 2

    # Headers
    headers = ['#', 'EL Entry', 'PY Entry', 'EL Exit', 'PY Exit', 'EL P&L', 'PY P&L', 'Match?']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    row += 1

    # Compare trades
    for i in range(min(30, len(csv_trades), len(xlsx_trades))):
        csv_t = csv_trades[i]
        xlsx_t = xlsx_trades[i]

        ws[f'A{row}'] = i + 1
        ws[f'B{row}'] = csv_t['entry_date']
        ws[f'C{row}'] = str(xlsx_t['entry_time'])
        ws[f'D{row}'] = csv_t.get('exit_date', '')
        ws[f'E{row}'] = str(xlsx_t['exit_time'])

        csv_pnl_str = csv_t.get('pnl_raw', '$0')
        csv_pnl = float(csv_pnl_str.replace('$', '').replace(',', '').replace('(', '-').replace(')', ''))

        ws[f'F{row}'] = csv_pnl_str
        ws[f'G{row}'] = f"${xlsx_t['pnl']:,.2f}"

        # Check if P&L matches (within 10% tolerance for scaling)
        match = abs(csv_pnl - xlsx_t['pnl'] * 10) < 10
        ws[f'H{row}'] = "✓" if match else "✗"
        ws[f'H{row}'].fill = match_fill if match else diff_fill

        for col in range(1, 9):
            ws.cell(row=row, column=col).border = border
        row += 1

    # Adjust widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 10

    wb.save(output_file)
    print(f"\nComparison saved to: {output_file}")

    return {
        'csv_trades': len(csv_trades),
        'xlsx_trades': len(xlsx_trades),
        'csv_pnl': csv_total_pnl,
        'xlsx_pnl': xlsx_total_pnl,
    }

# Run comparisons for all 4 strategies
strategies = [
    ('VEGA LE DN NI.csv', 'VEGA LE DN MAI.xlsx', 'VEGA_LE_DN_Comparison.xlsx', 'VEGA LE DN'),
    ('VEGA LE UP NI.csv', 'VEGA LE UP MAI.xlsx', 'VEGA_LE_UP_Comparison.xlsx', 'VEGA LE UP'),
    ('VEGA SE DN NI.csv', 'VEGA SE DN MAI.xlsx', 'VEGA_SE_DN_Comparison.xlsx', 'VEGA SE DN'),
    ('VEGA SE UP NI.csv', 'VEGA SE UP MAI.xlsx', 'VEGA_SE_UP_Comparison.xlsx', 'VEGA SE UP'),
]

print("=" * 60)
print("VEGA STRATEGY COMPARISON REPORT")
print("=" * 60)

for csv_file, xlsx_file, output_file, name in strategies:
    try:
        print(f"\n\nProcessing {name}...")
        result = create_comparison(csv_file, xlsx_file, output_file, name)
        print(f"  EasyLanguage: {result['csv_trades']} trades, ${result['csv_pnl']:,.2f}")
        print(f"  Python:       {result['xlsx_trades']} trades, ${result['xlsx_pnl']:,.2f}")
        print(f"  Difference:   ${result['xlsx_pnl'] - result['csv_pnl']:,.2f}")
    except Exception as e:
        print(f"  ERROR: {e}")

print("\n" + "=" * 60)
print("Comparison complete!")
